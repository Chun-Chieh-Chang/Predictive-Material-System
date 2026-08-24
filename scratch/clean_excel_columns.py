import openpyxl
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r'c:\Users\USER\Downloads\Project\Predictive-Material-System\新增資料夾\原料\塑膠原料清單明細.xlsx'
wb = openpyxl.load_workbook(file_path)

# Regex to match trailing (Code) like (A), (A01), (A0101), (B01), (I01), (H0101)
# Note: Allows optional spaces inside parens, e.g. ( I01 ) or ( A01 )
# Pattern: \s*\(\s*[A-Za-z]+[0-9]*\s*\)\s*$
def clean_trailing_code(text):
    if text is None:
        return text
    s = str(text).strip()
    if not s:
        return text
        
    # If the whole string is just (K) or (L) or (A)
    if re.match(r'^\s*\(\s*[A-Za-z0-9]+\s*\)\s*$', s):
        return re.sub(r'^\s*\(\s*([A-Za-z0-9]+)\s*\)\s*$', r'\1', s)
        
    # Match trailing (Code) where Code is letter followed by optional letters/digits, e.g. (A), (A01), (A0101), (I01)
    # Does NOT match (75-2567), (R1-15157), (台達), (TEKNIPLEX), (收縮膜), (新配方)
    cleaned = re.sub(r'\s*\(\s*[A-Za-z]+[0-9]*\s*\)\s*$', '', s)
    return cleaned.strip()

ws = wb['文件清單明細']
changes = []

for r in range(3, ws.max_row + 1):
    for c_idx, c_name in [(1, '原料種類'), (2, '廠商'), (3, '型號')]:
        cell = ws.cell(row=r, column=c_idx)
        val = cell.value
        if val is not None:
            new_val = clean_trailing_code(val)
            if new_val != val:
                changes.append((r, c_name, val, new_val))
                cell.value = new_val

print(f"Total cell changes: {len(changes)}")
# Print first 20 changes
print("\nSample changes:")
for ch in changes[:25]:
    print(f"Row {ch[0]:3d} | Col '{ch[1]}': {repr(ch[2])} -> {repr(ch[3])}")

# Save the workbook
wb.save(file_path)
print(f"\nSuccessfully saved updated workbook to: {file_path}")
