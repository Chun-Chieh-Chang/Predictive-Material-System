import json
import sys
import io
import os
import urllib.parse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

# Load json data
with open('scratch/color_powder_p1_raw.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

spans = data['spans']
links = data['links']

# 23 columns
col_defs = [
    {"name": "色粉料號", "sub": "Colorant Code", "x0": 90.0, "x1": 151.4},
    {"name": "成分", "sub": "Composition", "x0": 151.4, "x1": 259.8},
    {"name": "製造商", "sub": "Manufacturer", "x0": 259.8, "x1": 416.1},
    {"name": "原料名", "sub": "Raw Material Name", "x0": 416.1, "x1": 572.5},
    {"name": "比例", "sub": "SUM of Proportion", "x0": 572.5, "x1": 640.7},
    {"name": "MSDS / SDS", "sub": "MSDS / SDS", "x0": 640.7, "x1": 999.9},
    {"name": "TDS", "sub": "TDS", "x0": 999.9, "x1": 1087.4},
    {"name": "Animal origin", "sub": "Animal origin", "x0": 1087.4, "x1": 1185.7},
    {"name": "Phthalates", "sub": "Phthalates", "x0": 1185.7, "x1": 1284.0},
    {"name": "REACH SVHC", "sub": "REACH SVHC", "x0": 1284.0, "x1": 1382.3},
    {"name": "REACH AnnexXIV", "sub": "REACH AnnexXIV", "x0": 1382.3, "x1": 1480.7},
    {"name": "REACH AnnexXVII", "sub": "REACH AnnexXVII", "x0": 1480.7, "x1": 1579.0},
    {"name": "CLP", "sub": "CLP", "x0": 1579.0, "x1": 1677.3},
    {"name": "RoHS", "sub": "RoHS", "x0": 1677.3, "x1": 1775.6},
    {"name": "LatexFree", "sub": "LatexFree", "x0": 1775.6, "x1": 1874.0},
    {"name": "FDA21CFR", "sub": "FDA21CFR", "x0": 1874.0, "x1": 1972.3},
    {"name": "PFAS", "sub": "PFAS", "x0": 1972.3, "x1": 2070.6},
    {"name": "BPA", "sub": "BPA", "x0": 2070.6, "x1": 2168.9},
    {"name": "California Proposition 65", "sub": "Proposition 65", "x0": 2168.9, "x1": 2267.2},
    {"name": "Conflict Minerals", "sub": "Conflict Minerals", "x0": 2267.2, "x1": 2365.6},
    {"name": "Nanomaterial", "sub": "Nanomaterial", "x0": 2365.6, "x1": 2463.9},
    {"name": "POPs", "sub": "POPs", "x0": 2463.9, "x1": 2562.2},
    {"name": "MDRCMR", "sub": "MDRCMR", "x0": 2562.2, "x1": 2660.5},
]

row_y = [
    197.1, 215.6, 234.2, 252.8, 271.4, 289.9, 308.5, 327.1, 345.7, 364.3, 
    382.9, 401.4, 420.0, 438.6, 457.2, 475.8, 494.3, 512.9, 531.5, 550.1, 
    568.7, 587.2, 605.8, 624.4, 643.0, 661.6, 680.1, 698.7, 717.3, 735.9, 
    754.5, 773.0, 791.6, 810.2, 826.5, 842.7, 859.0, 875.2
]

# Colorant code grouping bounds
group_bounds = [
    (197.1, 215.6, ""),
    (215.6, 252.8, "73016"),
    (252.8, 289.9, "66821"),
    (289.9, 327.1, "77504"),
    (327.1, 345.7, "90682"),
    (345.7, 382.9, "66923"),
    (382.9, 420.0, "78392"),
    (420.0, 457.2, "77507"),
    (457.2, 494.3, "62310"),
    (494.3, 531.5, "73081"),
    (531.5, 568.7, "78346"),
    (568.7, 605.8, "61335"),
    (605.8, 661.6, "78360"),
    (661.6, 680.1, "CL2220"),
    (680.1, 717.3, "66914"),
    (717.3, 773.0, "77515"),
    (773.0, 810.2, "66950A"),
    (810.2, 842.7, "77493"),
    (842.7, 859.0, "同原料名"),
    (859.0, 875.2, "8905"),
]

all_rows = []
for r_idx in range(len(row_y)-1):
    yt, yb = row_y[r_idx], row_y[r_idx+1]
    
    code = ""
    for gb in group_bounds:
        if (yt + yb)/2 >= gb[0] - 2.0 and (yt + yb)/2 <= gb[1] + 2.0:
            code = gb[2]
            break
            
    row_dict = {"色粉料號": code}
    row_links = {}
    
    for c in col_defs[1:]:
        cname = c['name']
        c_spans = []
        for s in spans:
            if s['bbox'][1] >= 180:
                xc = (s['bbox'][0] + s['bbox'][2]) / 2
                yc = (s['bbox'][1] + s['bbox'][3]) / 2
                if yt - 2.0 <= yc < yb + 2.0 and c['x0'] - 2.0 <= xc < c['x1'] + 2.0:
                    c_spans.append(s)
                    
        c_spans = sorted(c_spans, key=lambda s: s['bbox'][0])
        # Clean text
        cell_text = " ".join([s['text'] for s in c_spans])
        # Normalize underscores like 2018-03-25_ 中 -> 2018-03-25 中
        cell_text = cell_text.replace('_ ', ' ').replace('_', ' ')
        row_dict[cname] = cell_text
        
        # Check links
        c_links = []
        for lk in links:
            l_xc = (lk['bbox'][0] + lk['bbox'][2]) / 2
            l_yc = (lk['bbox'][1] + lk['bbox'][3]) / 2
            if yt - 2.0 <= l_yc < yb + 2.0 and c['x0'] - 2.0 <= l_xc < c['x1'] + 2.0:
                # Find overlapping span text
                lk_bbox = lk['bbox']
                overlap = [s['text'].replace('_', ' ') for s in spans if not (s['bbox'][2] < lk_bbox[0] or s['bbox'][0] > lk_bbox[2] or s['bbox'][3] < lk_bbox[1] or s['bbox'][1] > lk_bbox[3])]
                lk['label'] = " ".join(overlap).strip()
                c_links.append(lk)
        if c_links:
            row_links[cname] = c_links
            
    all_rows.append({
        'row_idx': r_idx,
        'code': code,
        'cells': row_dict,
        'links': row_links
    })

# Build Excel
wb = openpyxl.Workbook()
ws_main = wb.active
ws_main.title = "色粉資料表"

# Colors & Fonts
NAVY_HEADER = "1F497D"
NAVY_SUBHEADER = "2E75B6"
WHITE = "FFFFFF"
BORDER_GRAY = "D9D9D9"
LINK_BLUE = "0563C1"

font_title = Font(name="微軟正黑體", size=14, bold=True, color="1F497D")
font_meta = Font(name="微軟正黑體", size=10, italic=True, color="595959")
font_h1 = Font(name="微軟正黑體", size=10, bold=True, color=WHITE)
font_h2 = Font(name="微軟正黑體", size=9, bold=True, color=WHITE)
font_data = Font(name="微軟正黑體", size=9, color="000000")
font_link = Font(name="微軟正黑體", size=9, color=LINK_BLUE, underline="single")
font_x = Font(name="微軟正黑體", size=9, color="808080")

fill_h1 = PatternFill(start_color=NAVY_HEADER, end_color=NAVY_HEADER, fill_type="solid")
fill_h2 = PatternFill(start_color=NAVY_SUBHEADER, end_color=NAVY_SUBHEADER, fill_type="solid")
fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

thin_side = Side(border_style="thin", color=BORDER_GRAY)
thick_bottom = Side(border_style="medium", color=NAVY_HEADER)
border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
border_header1 = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
border_header2 = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thick_bottom)

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Title & Metadata
ws_main.merge_cells("A1:W1")
ws_main["A1"] = "色粉資料表 (Colorant Formulation & Technical Data Sheet)"
ws_main["A1"].font = font_title
ws_main["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws_main.row_dimensions[1].height = 25

ws_main.merge_cells("A2:W2")
ws_main["A2"] = "修改日期：2026-07-14  |  資料來源：色粉資料表.pdf  |  點擊各檢驗文件日期/編號可直接開啟原廠 PDF 報告"
ws_main["A2"].font = font_meta
ws_main["A2"].alignment = Alignment(horizontal="left", vertical="center")
ws_main.row_dimensions[2].height = 18

# Header Rows: Row 3 & 4
ws_main.row_dimensions[3].height = 22
ws_main.row_dimensions[4].height = 20

for col_idx, c in enumerate(col_defs, start=1):
    cell1 = ws_main.cell(row=3, column=col_idx, value=c['name'])
    cell1.font = font_h1
    cell1.fill = fill_h1
    cell1.alignment = align_center
    cell1.border = border_header1
    
    cell2 = ws_main.cell(row=4, column=col_idx, value=c['sub'])
    cell2.font = font_h2
    cell2.fill = fill_h2
    cell2.alignment = align_center
    cell2.border = border_header2

# Insert Data Rows (starting from row 5)
start_row = 5
for r_idx, r in enumerate(all_rows):
    current_row = start_row + r_idx
    ws_main.row_dimensions[current_row].height = 24
    
    for col_idx, c in enumerate(col_defs, start=1):
        cname = c['name']
        cell_val = r['cells'].get(cname, '')
        cell = ws_main.cell(row=current_row, column=col_idx)
        cell.border = border_cell
        
        # Zebra
        if r_idx % 2 == 1:
            cell.fill = fill_zebra
            
        if col_idx in [1, 2, 3, 4]:
            cell.alignment = align_left
        else:
            cell.alignment = align_center
            
        cell.value = cell_val
        
        # Links
        c_links = r['links'].get(cname, [])
        if not cell_val:
            cell.font = font_data
        elif c_links and len(c_links) == 1:
            cell.font = font_link
            cell.hyperlink = c_links[0]['uri']
            cell.comment = openpyxl.comments.Comment(f"原廠文件連結:\n{c_links[0]['uri']}", "系統")
        elif c_links and len(c_links) > 1:
            cell.font = font_link
            cell.hyperlink = c_links[0]['uri']
            comment_text = "包含多個文件連結:\n" + "\n".join([f"{idx+1}. {lk.get('label', '')}: {lk['uri']}" for idx, lk in enumerate(c_links)])
            cell.comment = openpyxl.comments.Comment(comment_text, "系統")
        else:
            cell.font = font_data

# Adjust column widths
for col_idx, c in enumerate(col_defs, start=1):
    col_letter = get_column_letter(col_idx)
    if col_idx == 1: ws_main.column_dimensions[col_letter].width = 14
    elif col_idx == 2: ws_main.column_dimensions[col_letter].width = 22
    elif col_idx == 3: ws_main.column_dimensions[col_letter].width = 30
    elif col_idx == 4: ws_main.column_dimensions[col_letter].width = 32
    elif col_idx == 5: ws_main.column_dimensions[col_letter].width = 10
    elif col_idx == 6: ws_main.column_dimensions[col_letter].width = 38 # MSDS / SDS
    else:
        ws_main.column_dimensions[col_letter].width = 16

# ==========================================
# Sheet 2: 色粉文件明細 (Flat Catalog View)
# ==========================================
ws_detail = wb.create_sheet(title="色粉文件清單明細")

detail_headers = [
    ("項次", 8),
    ("色粉料號", 14),
    ("成分", 22),
    ("製造商", 28),
    ("原料名", 30),
    ("比例", 10),
    ("檢驗 / 項目類別", 24),
    ("文件說明 / 日期", 26),
    ("檔案名稱", 40),
    ("原廠文件下載連結 (點擊開啟)", 65)
]

ws_detail.row_dimensions[1].height = 25
ws_detail.merge_cells("A1:J1")
ws_detail["A1"] = "色粉原料 - 原廠檢驗證明與技術文件清單 (附直接下載連結)"
ws_detail["A1"].font = font_title
ws_detail["A1"].alignment = Alignment(horizontal="left", vertical="center")

ws_detail.row_dimensions[2].height = 20
for c_idx, (h_name, w) in enumerate(detail_headers, start=1):
    col_letter = get_column_letter(c_idx)
    ws_detail.column_dimensions[col_letter].width = w
    cell = ws_detail.cell(row=2, column=c_idx, value=h_name)
    cell.font = font_h1
    cell.fill = fill_h1
    cell.alignment = align_center
    cell.border = border_header1

detail_row_idx = 3
seq_num = 1

for r in all_rows:
    code = r['code']
    comp = r['cells']['成分']
    mfg = r['cells']['製造商']
    raw_name = r['cells']['原料名']
    ratio = r['cells']['比例']
    
    for c in col_defs[5:]: # columns with certificates / MSDS
        cname = c['name']
        col_header = f"{c['name']} ({c['sub']})" if c['sub'] != c['name'] else c['name']
        cell_val = r['cells'].get(cname, '')
        cell_links = r['links'].get(cname, [])
        
        if not cell_val:
            continue
            
        if cell_links:
            for lk in cell_links:
                uri = lk['uri']
                decoded_uri = urllib.parse.unquote(uri)
                fname = os.path.basename(decoded_uri)
                label = lk.get('label', cell_val)
                if not label: label = cell_val
                
                ws_detail.row_dimensions[detail_row_idx].height = 20
                ws_detail.cell(row=detail_row_idx, column=1, value=seq_num).alignment = align_center
                ws_detail.cell(row=detail_row_idx, column=2, value=code).alignment = align_center
                ws_detail.cell(row=detail_row_idx, column=3, value=comp).alignment = align_left
                ws_detail.cell(row=detail_row_idx, column=4, value=mfg).alignment = align_left
                ws_detail.cell(row=detail_row_idx, column=5, value=raw_name).alignment = align_left
                ws_detail.cell(row=detail_row_idx, column=6, value=ratio).alignment = align_center
                ws_detail.cell(row=detail_row_idx, column=7, value=col_header).alignment = align_left
                ws_detail.cell(row=detail_row_idx, column=8, value=label).alignment = align_left
                ws_detail.cell(row=detail_row_idx, column=9, value=fname).alignment = align_left
                
                link_cell = ws_detail.cell(row=detail_row_idx, column=10, value=decoded_uri)
                link_cell.font = font_link
                link_cell.alignment = align_left
                link_cell.hyperlink = uri
                
                for c_i in range(1, 11):
                    ws_detail.cell(row=detail_row_idx, column=c_i).border = border_cell
                    if (seq_num % 2) == 0:
                        ws_detail.cell(row=detail_row_idx, column=c_i).fill = fill_zebra
                    if c_i != 10:
                        ws_detail.cell(row=detail_row_idx, column=c_i).font = font_data
                        
                detail_row_idx += 1
                seq_num += 1

# ==========================================
# Sheet 3: 統計與摘要 (Summary)
# ==========================================
ws_summary = wb.create_sheet(title="統計與摘要")

ws_summary.merge_cells("A1:D1")
ws_summary["A1"] = "色粉資料表 - 統計與摘要"
ws_summary["A1"].font = font_title
ws_summary.row_dimensions[1].height = 25

summary_info = [
    ("資料版本日期", "2026-07-14"),
    ("色粉料號配方數", "20 款配方/規格 (73016, 66821, 77504, 90682, 66923, 78392, 77507, 62310, 73081, 78346, 61335, 78360, CL2220, 66914, 77515, 66950A, 77493, 同原料名, 8905 等)"),
    ("成分原料總項數", f"{len(all_rows)} 項成分配方"),
    ("代表製造廠牌", "Kronos International, Clariant, DCL / DCL2, BASF, LanXESS, Anthraquinone dyestuff, Daiichi Kasei Kogyo Co., Ltd."),
    ("有效原廠技術/證明文件總數", f"{seq_num - 1} 份"),
]

for idx, (label, val) in enumerate(summary_info, start=3):
    ws_summary.row_dimensions[idx].height = 22
    c1 = ws_summary.cell(row=idx, column=1, value=label)
    c1.font = font_h2
    c1.fill = fill_h2
    c1.border = border_cell
    c1.alignment = align_left
    
    ws_summary.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=4)
    c2 = ws_summary.cell(row=idx, column=2, value=val)
    c2.font = font_data
    c2.border = border_cell
    c2.alignment = align_left

ws_summary.column_dimensions["A"].width = 25
ws_summary.column_dimensions["B"].width = 30
ws_summary.column_dimensions["C"].width = 30
ws_summary.column_dimensions["D"].width = 30

# Save files
out_file1 = r"c:\Users\USER\Downloads\Project\Predictive-Material-System\新增資料夾\原料\色粉資料表.xlsx"
out_file2 = r"c:\Users\USER\Downloads\Project\Predictive-Material-System\色粉資料表.xlsx"

wb.save(out_file1)
wb.save(out_file2)

print(f"Exported successfully to:\n1. {out_file1}\n2. {out_file2}")
print(f"Total detail document records: {seq_num - 1}")
