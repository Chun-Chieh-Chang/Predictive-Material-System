import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r'c:\Users\USER\Downloads\Project\Predictive-Material-System\新增資料夾\原料\塑膠原料清單明細.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print(f"Original total rows: {ws.max_row}")

# Read all data rows (row 3 to max_row)
rows_data = []
for r in range(3, ws.max_row + 1):
    vals = [ws.cell(row=r, column=c).value for c in range(1, 5)]
    if any(vals):
        rows_data.append(vals)

# Deduplicate while preserving original order
seen = set()
unique_rows = []
for r in rows_data:
    t = tuple(r)
    if t not in seen:
        seen.add(t)
        unique_rows.append(r)

print(f"Deduplicated count: {len(unique_rows)} unique rows")

# Styles
NAVY_HEADER = "1F497D"
WHITE = "FFFFFF"
BORDER_GRAY = "D9D9D9"

font_title = Font(name="微軟正黑體", size=14, bold=True, color="1F497D")
font_header = Font(name="微軟正黑體", size=11, bold=True, color=WHITE)
font_data = Font(name="微軟正黑體", size=10, color="000000")

fill_header = PatternFill(start_color=NAVY_HEADER, end_color=NAVY_HEADER, fill_type="solid")
fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

thin_side = Side(border_style="thin", color=BORDER_GRAY)
border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Clear entire worksheet content to recreate cleanly
wb.remove(ws)
ws_new = wb.create_sheet(title="文件清單明細", index=0)

# Title Row (Row 1)
ws_new.merge_cells("A1:D1")
ws_new["A1"] = "塑膠原料 - 清單明細"
ws_new["A1"].font = font_title
ws_new["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws_new.row_dimensions[1].height = 28

# Header Row (Row 2)
headers = ['原料種類', '廠商', '型號', '顏色']
ws_new.row_dimensions[2].height = 24
for col_idx, h in enumerate(headers, start=1):
    cell = ws_new.cell(row=2, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_cell

# Write unique data rows
for r_idx, row_vals in enumerate(unique_rows):
    current_row = 3 + r_idx
    ws_new.row_dimensions[current_row].height = 22
    
    for c_idx, val in enumerate(row_vals, start=1):
        cell = ws_new.cell(row=current_row, column=c_idx, value=val)
        cell.font = font_data
        cell.border = border_cell
        
        # Zebra striping
        if r_idx % 2 == 1:
            cell.fill = fill_zebra
        else:
            cell.fill = fill_white
            
        # Alignment
        if c_idx in [1, 2, 3]:
            cell.alignment = align_left
        else:
            cell.alignment = align_center

# Column widths
ws_new.column_dimensions['A'].width = 24  # 原料種類
ws_new.column_dimensions['B'].width = 36  # 廠商
ws_new.column_dimensions['C'].width = 48  # 型號
ws_new.column_dimensions['D'].width = 16  # 顏色

wb.save(file_path)
print(f"Successfully saved deduplicated file to: {file_path}")
print(f"New sheet max_row: {ws_new.max_row}, max_col: {ws_new.max_column}")
