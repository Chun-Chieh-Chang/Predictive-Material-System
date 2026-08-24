import openpyxl
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r'c:\Users\USER\Downloads\Project\Predictive-Material-System\新增資料夾\原料\塑膠原料清單明細.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

cat_vals = set()
mfg_vals = set()
model_vals = set()

for r in range(3, ws.max_row + 1):
    c1 = ws.cell(row=r, column=1).value
    c2 = ws.cell(row=r, column=2).value
    c3 = ws.cell(row=r, column=3).value
    if c1 is not None: cat_vals.add(str(c1))
    if c2 is not None: mfg_vals.add(str(c2))
    if c3 is not None: model_vals.add(str(c3))

def clean_trailing_code(text):
    if not text:
        return text
    s = str(text).strip()
    
    # We want to remove trailing parentheses containing code like (A), (A01), (A0101), (B), (B01), (I), (J), (K), (L), etc.
    # Note:
    # If text is '(K)', then what should it become? Let's check!
    # In category: 'ABS (A)' -> 'ABS'
    # 'Acrylic-Polycarbonate Alloys (I)' -> 'Acrylic-Polycarbonate Alloys'
    # 'Copolyester (H)' -> 'Copolyester'
    # 'PBT (F)' -> 'PBT'
    # 'PC (C)' -> 'PC'
    # 'PE (D)' -> 'PE'
    # 'PP (E)' -> 'PP'
    # 'PVC (B)' -> 'PVC'
    # 'SBC (J)' -> 'SBC'
    # 'TPE (G)' -> 'TPE'
    # '(K)' -> 'K' (or '(K)'? If it's just '(K)', removing outer parens makes it 'K')
    
    # In manufacturer:
    # 'Benison (本源興) (B05)' -> 'Benison (本源興)'
    # 'Bormed (E02)' -> 'Bormed'
    # 'COLORITE (TEKNIPLEX) (B02)' -> 'COLORITE (TEKNIPLEX)'
    # 'EASTMAN (H01)' -> 'EASTMAN'
    # 'ELINT HILLS (E03)' -> 'ELINT HILLS'
    # 'EVONIK (I01)' -> 'EVONIK'
    # 'Formerra (PolyOne) (B03)' -> 'Formerra (PolyOne)'
    # 'INEOS (A03)' -> 'INEOS'
    # 'INEOS (J01)' -> 'INEOS'
    # 'JIEH-MING (介明) (B06)' -> 'JIEH-MING (介明)'
    # 'LCY (李長榮) (E01)' -> 'LCY (李長榮)'
    # 'NAN YA(南亞) (B01)' -> 'NAN YA(南亞)'
    # 'SABIC (C02)' -> 'SABIC'
    # 'SABIC (F01)' -> 'SABIC'
    # 'Saint-Gobain Performance Plastics (G01)' -> 'Saint-Gobain Performance Plastics'
    # 'TAITA (台達) (A02)' -> 'TAITA (台達)'
    # 'TEKNOR APEX (B07)' -> 'TEKNOR APEX'
    # 'TORAY (A01)' -> 'TORAY'
    # 'USI (台聚) (D01)' -> 'USI (台聚)'
    # 'ZEON / Lucky Seal (K01)' -> 'ZEON / Lucky Seal'
    # 'axiall (B04)' -> 'axiall'
    # 'covestro (Bayer) (C01)' -> 'covestro (Bayer)'
    # 'lyondellbasell industries (D02)' -> 'lyondellbasell industries'
    
    # In model:
    # 'ABS TOYOLAC 900 352U (A0101)' -> 'ABS TOYOLAC 900 352U'
    # 'ABS Terlux® 2802 (A0303)' -> 'ABS Terlux® 2802'
    # 'ABS Terlux® 2812 (A0304)' -> 'ABS Terlux® 2812'
    # 'Alathon® HDPE M6580 (D0201)' -> 'Alathon® HDPE M6580'
    # 'C-Flex ® R70-028-000  (G0101)' -> 'C-Flex ® R70-028-000'
    # 'CYREX ® Acrylic-Polycarbonate Alloys (R1-1092)  (I0101)' -> 'CYREX ® Acrylic-Polycarbonate Alloys (R1-1092)'
    # 'Geon TM 161J (75-1861) (B0302)' -> 'Geon TM 161J (75-1861)'
    # 'HDPE UNITHENE LH606 (D0101)' -> 'HDPE UNITHENE LH606'
    # 'I-632 NT CLEAR 0053 (B0701)' -> 'I-632 NT CLEAR 0053'
    # 'K-Resin®  (J0101)' -> 'K-Resin®'
    # 'LDPE PAXOTHENE NA207-66 (D0102)' -> 'LDPE PAXOTHENE NA207-66'
    # 'Lustran® 348 NR (75-2567)  (A0302)' -> 'Lustran® 348 NR (75-2567)'
    # 'Lustran® 348 WT012002 (75-2568) (A0301)' -> 'Lustran® 348 WT012002 (75-2568)'
    # 'Nipol IR2200 (B膠)  (K0101)' -> 'Nipol IR2200 (B膠)'
    # 'NonP2 700 86A (B0601)' -> 'NonP2 700 86A'
    # 'P4G3Z-039 (R1-10002) (E0301)' -> 'P4G3Z-039 (R1-10002)'
    # 'PC LEXAN TM 144R (C0201)' -> 'PC LEXAN TM 144R'
    # 'PC MAKROLON 2458-550115 (C0103)' -> 'PC MAKROLON 2458-550115'
    # 'PC MAKROLON 2558-558882 (C0105)' -> 'PC MAKROLON 2558-558882'
    # 'PC MAKROLON RX1805-013771 (C0102)' -> 'PC MAKROLON RX1805-013771'
    # 'PC MAKROLON RX1805-451118 (C0101)' -> 'PC MAKROLON RX1805-451118'
    # 'PC MAKROLON RX2530-451118 (C0104)' -> 'PC MAKROLON RX2530-451118'
    # 'PP Bormed™ HD810MO (E0201)' -> 'PP Bormed™ HD810MO'
    # 'PP GLOBALENE 6331 (E0101)' -> 'PP GLOBALENE 6331'
    # 'PP GLOBALENE 8661 (新配方) (E0103)' -> 'PP GLOBALENE 8661 (新配方)'
    # 'PP GLOBALENE 8661A (舊配方) (E0102)' -> 'PP GLOBALENE 8661A (舊配方)'
    # 'PVC 3MSA044GXX003 (B0103)' -> 'PVC 3MSA044GXX003'
    # 'PVC 3MSA048P3X000 (B0104)' -> 'PVC 3MSA048P3X000'
    # 'PVC 3MSA055GXX003 (B0105)' -> 'PVC 3MSA055GXX003'
    # 'PVC 3MSA140P320P0 (B0106)' -> 'PVC 3MSA140P320P0'
    # 'PVC 3MTA002GXX001 (B0101)' -> 'PVC 3MTA002GXX001'
    # 'PVC 3MTA044GXX002 (B0102)' -> 'PVC 3MTA044GXX002'
    # 'PVC 7088G-015 (B0201)' -> 'PVC 7088G-015'
    # 'PVC 7477G-015 (B0202)' -> 'PVC 7477G-015'
    # 'PVC 8512G-015 (B0203)' -> 'PVC 8512G-015'
    # 'PVC 8577G-015 (B0204)' -> 'PVC 8577G-015'
    # 'PVC GEON M4910 TRANS 9494 (R1-15157) (B0301)' -> 'PVC GEON M4910 TRANS 9494 (R1-15157)'
    # 'PVC SHRINKABLE FILM (收縮膜) (B0501)' -> 'PVC SHRINKABLE FILM (收縮膜)'
    # 'Rigid PVC Compounds-Pellet All Colors (90-9634) (B0401)' -> 'Rigid PVC Compounds-Pellet All Colors (90-9634)'
    # 'TAITALAC 1000 (A0201)' -> 'TAITALAC 1000'
    # 'TAITALAC 1000 W-767 (A0202)' -> 'TAITALAC 1000 W-767'
    # 'TRITAN MX711 (R1-8328)  (H0101)' -> 'TRITAN MX711 (R1-8328)'
    # 'VALOX TM HX420HP (R1-1176) (F0101)' -> 'VALOX TM HX420HP (R1-1176)'
    # 'VALOX TM RESIN 325-1001 (F0102)' -> 'VALOX TM RESIN 325-1001'

    # If exact match for single code in parentheses like '(K)' or '(L)'
    if re.match(r'^\s*\(\s*[A-Za-z0-9]+\s*\)\s*$', s):
        # Could be 'K' or '(K)' depending on context, or if user said: "移除儲存格內尾綴括號內的英文+數字組合，如(A)、(A01)、(A0101)"
        # If it's category '(K)', stripping the parens leaves 'K'
        return re.sub(r'^\s*\(\s*([A-Za-z0-9]+)\s*\)\s*$', r'\1', s)
        
    # Match trailing (Code) where Code is alphanumeric like [A-Za-z0-9]+ (or with spaces like 'I 01', 'J 01')
    cleaned = re.sub(r'\s*\(\s*[A-Za-z][A-Za-z0-9]*\s*\)\s*$', '', s)
    return cleaned.strip()

print("\n=== Test 原料種類 ===")
for v in sorted(cat_vals):
    print(f"{repr(v):<40} -> {repr(clean_trailing_code(v))}")

print("\n=== Test 廠商 ===")
for v in sorted(mfg_vals):
    print(f"{repr(v):<45} -> {repr(clean_trailing_code(v))}")

print("\n=== Test 型號 ===")
for v in sorted(model_vals):
    print(f"{repr(v):<60} -> {repr(clean_trailing_code(v))}")
