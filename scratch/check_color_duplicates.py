import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r'c:\Users\USER\Downloads\Project\Predictive-Material-System\新增資料夾\原料\色粉清單明細.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['色粉清單明細']

print(f"Total rows: {ws.max_row}, Total cols: {ws.max_column}")
print("Row 1 (Title):", [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)])
print("Row 2 (Headers):", [ws.cell(row=2, column=c).value for c in range(1, ws.max_column+1)])

# Let's inspect data rows (col 2..5: 色粉料號, 成分, 製造商, 原料名)
raw_rows = []
for r in range(3, ws.max_row + 1):
    # Col 1 is 项次 (1..466)
    # Col 2 is 色粉料號
    # Col 3 is 成分
    # Col 4 is 製造商
    # Col 5 is 原料名
    vals = [ws.cell(row=r, column=c).value for c in range(2, 6)]
    raw_rows.append(vals)

print(f"Total raw data rows: {len(raw_rows)}")

# Deduplicate based on (色粉料號, 成分, 製造商, 原料名)
seen = set()
unique_rows = []
for r in raw_rows:
    t = tuple(r)
    if t not in seen:
        seen.add(t)
        unique_rows.append(r)

print(f"Unique data rows: {len(unique_rows)}")
for idx, u in enumerate(unique_rows, 1):
    print(f"{idx:2d}: {u}")
