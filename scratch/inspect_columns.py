import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r'c:\Users\USER\Downloads\Project\Predictive-Material-System\新增資料夾\原料\塑膠原料清單明細.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print("Sheet Title:", ws.title)
print("Max row:", ws.max_row, "Max col:", ws.max_column)

# Print headers
headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
print("Headers in Row 2:", headers)

# Inspect distinct values in col 1, 2, 3
cat_vals = set()
mfg_vals = set()
model_vals = set()

for r in range(3, ws.max_row + 1):
    c1 = ws.cell(row=r, column=1).value
    c2 = ws.cell(row=r, column=2).value
    c3 = ws.cell(row=r, column=3).value
    if c1 is not None: cat_vals.add(str(c1))
    if c2 is not None: mfg_vals.add(str(c2))
    if c3 is not None: model_vals.add(str(c3))

print("\n--- Distinct 原料種類 (Cat) ---")
for v in sorted(cat_vals):
    print(f"  {repr(v)}")

print("\n--- Distinct 廠商 (Mfg) ---")
for v in sorted(mfg_vals):
    print(f"  {repr(v)}")

print("\n--- Distinct 型號 (Model) ---")
for v in sorted(model_vals):
    print(f"  {repr(v)}")
