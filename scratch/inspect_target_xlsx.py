import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r'c:\Users\USER\Downloads\Project\Predictive-Material-System\新增資料夾\原料\塑膠原料清單明細.xlsx'
wb = openpyxl.load_workbook(file_path)

print(f"File: {file_path}")
print(f"Sheet names: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n--- Sheet: {sheet_name} (rows: {ws.max_row}, cols: {ws.max_column}) ---")
    for r in range(1, min(15, ws.max_row + 1)):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(15, ws.max_column + 1))]
        print(f"Row {r:2d}: {row_vals}")
