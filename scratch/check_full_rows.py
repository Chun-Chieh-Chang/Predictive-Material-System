import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r'c:\Users\USER\Downloads\Project\Predictive-Material-System\新增資料夾\原料\塑膠原料清單明細.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print(f"Total rows in {file_path}: {ws.max_row}")
for r in range(1, 15):
    print(f"Row {r:2d}: col1={repr(ws.cell(row=r, column=1).value)}, col2={repr(ws.cell(row=r, column=2).value)}, col3={repr(ws.cell(row=r, column=3).value)}, col4={repr(ws.cell(row=r, column=4).value)}")

# Check last 10 rows
print("\nLast 10 rows:")
for r in range(max(1, ws.max_row - 10), ws.max_row + 1):
    print(f"Row {r:2d}: col1={repr(ws.cell(row=r, column=1).value)}, col2={repr(ws.cell(row=r, column=2).value)}, col3={repr(ws.cell(row=r, column=3).value)}, col4={repr(ws.cell(row=r, column=4).value)}")
