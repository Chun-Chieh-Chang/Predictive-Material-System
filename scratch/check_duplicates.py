import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r'c:\Users\USER\Downloads\Project\Predictive-Material-System\新增資料夾\原料\塑膠原料清單明細.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print(f"Sheet title: {ws.title}")
print(f"Total rows before deduplication: {ws.max_row}")
print(f"Row 1: {[ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]}")
print(f"Row 2: {[ws.cell(row=2, column=c).value for c in range(1, ws.max_column+1)]}")

rows_data = []
for r in range(3, ws.max_row + 1):
    vals = [ws.cell(row=r, column=c).value for c in range(1, 5)] # Col 1..4: 原料種類, 廠商, 型號, 顏色
    rows_data.append(vals)

print(f"Total data rows: {len(rows_data)}")

# Count unique rows
seen = set()
unique_rows = []
for r in rows_data:
    t = tuple(r)
    if t not in seen:
        seen.add(t)
        unique_rows.append(r)

print(f"Unique data rows: {len(unique_rows)}")
print("\nUnique rows preview:")
for idx, r in enumerate(unique_rows, 1):
    print(f"{idx:2d}: {r}")
