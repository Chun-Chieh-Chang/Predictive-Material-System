import json
import sys
import io
import os
import shutil
import urllib.parse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

# Load json data
with open('scratch/pdf_raw.json', 'r', encoding='utf-8') as f:
    raw_data = json.load(f)

with open('scratch/spans.json', 'r', encoding='utf-8') as f:
    spans = json.load(f)

with open('scratch/links.json', 'r', encoding='utf-8') as f:
    links = json.load(f)

col_defs = [
    {"name": "原料種類", "sub": "Category", "x0": 14.0, "x1": 36.5},
    {"name": "廠商", "sub": "Manufacturer", "x0": 36.5, "x1": 65.5},
    {"name": "型號", "sub": "Grade/Model", "x0": 65.5, "x1": 126.2},
    {"name": "顏色", "sub": "Color", "x0": 126.2, "x1": 138.7},
    {"name": "物理性資料", "sub": "TECHNICAL DATA SHEET", "x0": 138.7, "x1": 160.8},
    {"name": "物質安全資料", "sub": "MSDS / SDS", "x0": 160.8, "x1": 186.7},
    {"name": "生物相容性資料", "sub": "BIO (ISO 10993)", "x0": 186.7, "x1": 226.5},
    {"name": "動物性成分資料", "sub": "Animal Origin (BSE/TSE)", "x0": 226.5, "x1": 253.4},
    {"name": "塑化劑", "sub": "Phthalates", "x0": 253.4, "x1": 277.9},
    {"name": "REACH", "sub": "SVHC", "x0": 277.9, "x1": 306.5},
    {"name": "REACH", "sub": "Annex XIV", "x0": 306.5, "x1": 332.4},
    {"name": "REACH", "sub": "Annex XVII", "x0": 332.4, "x1": 358.3},
    {"name": "CLP", "sub": "CLP", "x0": 358.3, "x1": 385.9},
    {"name": "RoHS", "sub": "RoHS", "x0": 385.9, "x1": 409.9},
    {"name": "LatexFree", "sub": "LatexFree", "x0": 409.9, "x1": 428.4},
    {"name": "FDA21CFR", "sub": "FDA21CFR", "x0": 428.4, "x1": 446.9},
    {"name": "PFAS", "sub": "PFAS", "x0": 446.9, "x1": 473.3},
    {"name": "雙酚A", "sub": "BPA", "x0": 473.3, "x1": 497.5},
    {"name": "美國藥典(第六級)", "sub": "USP(CLASS VI)", "x0": 497.5, "x1": 526.8},
    {"name": "加州 65 號法案", "sub": "California Proposition 65", "x0": 526.8, "x1": 552.5},
    {"name": "衝突礦物", "sub": "Conflict Minerals free", "x0": 552.5, "x1": 575.0},
    {"name": "奈米材料", "sub": "Nanomaterial", "x0": 575.0, "x1": 601.9},
    {"name": "成型技術文件", "sub": "Molding Tech Doc", "x0": 601.9, "x1": 631.2},
    {"name": "POPs Regulation", "sub": "(EU) 2019/1021", "x0": 631.2, "x1": 660.5},
    {"name": "MDR CMR / ATPxx", "sub": "MDR CMR / ATPxx", "x0": 660.5, "x1": 689.7},
    {"name": "其它", "sub": "Other", "x0": 689.7, "x1": 813.1},
]

# Row Y divider lines
h_y = []
for l in raw_data['lines']:
    if l['type'] == 'l':
        x1, y1 = l['p1']
        x2, y2 = l['p2']
        if abs(y1 - y2) < 0.5 and min(x1, x2) <= 70 and max(x1, x2) >= 120 and y1 >= 45:
            h_y.append((y1+y2)/2)
    elif l['type'] == 're':
        x0, y0, x1, y1 = l['rect']
        if x0 <= 70 and x1 >= 120 and y0 >= 45:
            h_y.append(y0)
            h_y.append(y1)

h_y = sorted(list(set([round(y, 1) for y in h_y])))
merged_y = []
for y in h_y:
    if not merged_y or y - merged_y[-1] > 1.0:
        merged_y.append(y)

cat_bounds = [
    (47.3, 110.1, "ABS (A)"),
    (110.1, 283.4, "PVC (B)"),
    (283.4, 323.7, "PC (C)"),
    (323.7, 347.5, "PE (D)"),
    (347.5, 387.8, "PP (E)"),
    (387.8, 401.3, "PBT (F)"),
    (401.3, 420.5, "TPE (G)"),
    (420.5, 430.3, "Copolyester (H)"),
    (430.3, 451.7, "Acrylic-Polycarbonate Alloys (I)"),
    (451.7, 461.5, "SBC (J)"),
    (461.5, 476.4, "(K)"),
    (476.4, 503.3, "(L)"),
]

mfg_bounds = [
    (47.3, 59.3, "TORAY (A01)"),
    (59.3, 79.4, "TAITA (台達) (A02)"),
    (79.4, 110.1, "INEOS (A03)"),
    (110.1, 190.8, "NAN YA(南亞) (B01)"),
    (190.8, 217.7, "COLORITE (TEKNIPLEX) (B02)"),
    (217.7, 235.4, "Formerra (PolyOne) (B03)"),
    (235.4, 249.8, "axiall (B04)"),
    (249.8, 259.7, "Benison (本源興) (B05)"),
    (259.7, 270.0, "JIEH-MING (介明) (B06)"),
    (270.0, 283.4, "TEKNOR APEX (B07)"),
    (283.4, 317.0, "covestro (Bayer) (C01)"),
    (317.0, 323.7, "SABIC (C02)"),
    (323.7, 337.2, "USI (台聚) (D01)"),
    (337.2, 347.5, "lyondellbasell industries (D02)"),
    (347.5, 367.7, "LCY (李長榮) (E01)"),
    (367.7, 381.1, "Bormed (E02)"),
    (381.1, 387.8, "ELINT HILLS (E03)"),
    (387.8, 401.3, "SABIC (F01)"),
    (401.3, 420.5, "Saint-Gobain Performance Plastics (G01)"),
    (420.5, 430.3, "EASTMAN (H01)"),
    (430.3, 451.7, "EVONIK (I01)"),
    (451.7, 461.5, "INEOS (J01)"),
    (461.5, 476.4, "ZEON / Lucky Seal (K01)"),
    (476.4, 503.3, "GVS (L01)"),
]

def clean_model_name(name):
    name = name.strip()
    if name.endswith('(C0101'): name = name + ')'
    elif name.endswith('(C0102'): name = name + ')'
    elif name.endswith('(C0104'): name = name + ')'
    elif name.startswith('® CYREX'): name = 'CYREX® ' + name.replace('® CYREX', '').strip()
    elif name.startswith('® C-Flex'): name = 'C-Flex® ' + name.replace('® C-Flex', '').strip()
    elif name.startswith('TM VALOX'): name = 'VALOX™ ' + name.replace('TM VALOX', '').strip()
    return name

# Parse all 45 material rows
rows_data = []
all_detail_records = []

for i in range(len(merged_y)-2):
    yt = merged_y[i]
    yb = merged_y[i+1]
    
    cat = ""
    for cb in cat_bounds:
        if (yt + yb)/2 >= cb[0] - 0.5 and (yt + yb)/2 <= cb[1] + 0.5:
            cat = cb[2]
            break
            
    mfg = ""
    for mb in mfg_bounds:
        if (yt + yb)/2 >= mb[0] - 0.5 and (yt + yb)/2 <= mb[1] + 0.5:
            mfg = mb[2]
            break
            
    row_dict = {"原料種類": cat, "廠商": mfg}
    row_links = {}
    
    for c in col_defs[2:]:
        cname = c['name']
        col_header_full = f"{c['name']} ({c['sub']})" if c['sub'] != c['name'] else c['name']
        
        # find spans
        c_spans = []
        for s in spans:
            if s['bbox'][1] >= 47:
                xc = (s['bbox'][0] + s['bbox'][2]) / 2
                yc = (s['bbox'][1] + s['bbox'][3]) / 2
                if yt - 0.8 <= yc < yb + 0.8 and c['x0'] - 0.5 <= xc < c['x1'] + 0.5:
                    c_spans.append(s)
        
        # Group spans by line
        lines_in_cell = []
        c_spans_sorted = sorted(c_spans, key=lambda s: (s['bbox'][1], s['bbox'][0]))
        for s in c_spans_sorted:
            if not lines_in_cell:
                lines_in_cell.append([s])
            else:
                last_line_y = sum([x['bbox'][1] for x in lines_in_cell[-1]]) / len(lines_in_cell[-1])
                if abs(s['bbox'][1] - last_line_y) < 2.5:
                    lines_in_cell[-1].append(s)
                else:
                    lines_in_cell.append([s])
        
        line_texts = []
        for l in lines_in_cell:
            l = sorted(l, key=lambda s: s['bbox'][0])
            line_texts.append(" ".join([s['text'] for s in l]))
        
        cell_text = "\n".join(line_texts)
        if cname == '型號':
            cell_text = clean_model_name(cell_text)
        row_dict[cname] = cell_text
        
        # Find links in this cell
        c_links = []
        for lk in links:
            l_xc = (lk['bbox'][0] + lk['bbox'][2]) / 2
            l_yc = (lk['bbox'][1] + lk['bbox'][3]) / 2
            if yt - 0.8 <= l_yc < yb + 0.8 and c['x0'] - 0.5 <= l_xc < c['x1'] + 0.5:
                c_links.append(lk)
        
        if c_links:
            row_links[cname] = c_links
            
    rows_data.append({
        'row_idx': i,
        'cat': cat,
        'mfg': mfg,
        'cells': row_dict,
        'links': row_links
    })

# Build Excel Workbook
wb = openpyxl.Workbook()
ws_main = wb.active
ws_main.title = "塑膠原料資料表"

# Color tokens
NAVY_HEADER = "1F497D"
NAVY_SUBHEADER = "2E75B6"
LIGHT_BLUE_FILL = "EDF2F8"
WHITE = "FFFFFF"
GRAY_TEXT = "595959"
BORDER_GRAY = "D9D9D9"
LINK_BLUE = "0563C1"

font_title = Font(name="微軟正黑體", size=14, bold=True, color="1F497D")
font_meta = Font(name="微軟正黑體", size=10, italic=True, color="595959")
font_h1 = Font(name="微軟正黑體", size=10, bold=True, color=WHITE)
font_h2 = Font(name="微軟正黑體", size=9, bold=True, color=WHITE)
font_data = Font(name="微軟正黑體", size=9, color="000000")
font_link = Font(name="微軟正黑體", size=9, color=LINK_BLUE, underline="single")
font_x = Font(name="微軟正黑體", size=9, color="808080")

fill_h1 = PatternFill(start_color=NAVY_HEADER, end_color=NAVY_HEADER, fill_type="solid")
fill_h2 = PatternFill(start_color=NAVY_SUBHEADER, end_color=NAVY_SUBHEADER, fill_type="solid")
fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

thin_side = Side(border_style="thin", color=BORDER_GRAY)
thick_bottom = Side(border_style="medium", color=NAVY_HEADER)
border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
border_header1 = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
border_header2 = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thick_bottom)

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Title and metadata
ws_main.merge_cells("A1:Z1")
ws_main["A1"] = "塑膠原料資料表 (Plastic Raw Material Data Sheet)"
ws_main["A1"].font = font_title
ws_main["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws_main.row_dimensions[1].height = 25

ws_main.merge_cells("A2:Z2")
ws_main["A2"] = "修改日期：2026-08-10  |  資料來源：原料資料庫 / 塑膠原料資料表.pdf  |  點擊各文件編號可直接開啟/下載原廠報告 (PDF)"
ws_main["A2"].font = font_meta
ws_main["A2"].alignment = Alignment(horizontal="left", vertical="center")
ws_main.row_dimensions[2].height = 18

# Header Rows: Row 3 (Primary Header) and Row 4 (Secondary Header)
ws_main.row_dimensions[3].height = 22
ws_main.row_dimensions[4].height = 20

for col_idx, c in enumerate(col_defs, start=1):
    cell1 = ws_main.cell(row=3, column=col_idx, value=c['name'])
    cell1.font = font_h1
    cell1.fill = fill_h1
    cell1.alignment = align_center
    cell1.border = border_header1
    
    cell2 = ws_main.cell(row=4, column=col_idx, value=c['sub'])
    cell2.font = font_h2
    cell2.fill = fill_h2
    cell2.alignment = align_center
    cell2.border = border_header2

# Insert Data Rows (starting from Row 5)
start_row = 5
for r_idx, r in enumerate(rows_data):
    current_row = start_row + r_idx
    ws_main.row_dimensions[current_row].height = 28 if '\n' in r['cells']['型號'] else 22
    
    for col_idx, c in enumerate(col_defs, start=1):
        cname = c['name']
        cell_val = r['cells'].get(cname, '')
        cell = ws_main.cell(row=current_row, column=col_idx)
        cell.border = border_cell
        
        # Zebra coloring
        if r_idx % 2 == 1:
            cell.fill = fill_zebra
            
        if col_idx in [1, 2, 3]: # Category, Manufacturer, Model
            cell.alignment = align_left
        else:
            cell.alignment = align_center
            
        cell.value = cell_val
        
        # Link styling
        cell_links = r['links'].get(cname, [])
        if cell_val == 'X':
            cell.font = font_x
        elif cell_links and len(cell_links) == 1:
            cell.font = font_link
            cell.hyperlink = cell_links[0]['uri']
            cell.comment = openpyxl.comments.Comment(f"文件網址:\n{cell_links[0]['uri']}", "系統")
        elif cell_links and len(cell_links) > 1:
            cell.font = font_link
            # Set hyperlink to first link, and add tooltip comment for all links
            cell.hyperlink = cell_links[0]['uri']
            comment_text = "包含多個文件連結:\n" + "\n".join([f"{idx+1}. {lk['uri']}" for idx, lk in enumerate(cell_links)])
            cell.comment = openpyxl.comments.Comment(comment_text, "系統")
        else:
            cell.font = font_data

# Footer Row (化學物質耐受性參考表)
footer_row = start_row + len(rows_data)
ws_main.row_dimensions[footer_row].height = 22
ws_main.merge_cells(f"A{footer_row}:B{footer_row}")
cell_ft = ws_main.cell(row=footer_row, column=1, value="化學物質耐受性參考表")
cell_ft.font = font_link
cell_ft.alignment = align_center
cell_ft.hyperlink = "http://60.251.196.100/tw/%E5%8E%9F%E6%96%99%E8%B3%87%E6%96%99/ChemRes(%E5%8C%96%E5%AD%B8%E7%89%A9%E8%B3%AA%E8%80%90%E5%8F%97%E6%80%A7%E5%8F%83%E8%80%83%E8%A1%A8).pdf"
cell_ft.border = border_cell
ws_main.cell(row=footer_row, column=2).border = border_cell

for c_idx in range(3, len(col_defs)+1):
    c_empty = ws_main.cell(row=footer_row, column=c_idx, value="")
    c_empty.border = border_cell

# Adjust column widths for Sheet 1
for col_idx, c in enumerate(col_defs, start=1):
    col_letter = get_column_letter(col_idx)
    if col_idx == 1: ws_main.column_dimensions[col_letter].width = 16
    elif col_idx == 2: ws_main.column_dimensions[col_letter].width = 24
    elif col_idx == 3: ws_main.column_dimensions[col_letter].width = 38
    elif col_idx == 4: ws_main.column_dimensions[col_letter].width = 12
    elif col_idx == 26: ws_main.column_dimensions[col_letter].width = 35 # 其它
    else:
        ws_main.column_dimensions[col_letter].width = 16

# ==========================================
# Sheet 2: 文件清單明細 (包含所有可下載檔案清單)
# ==========================================
ws_detail = wb.create_sheet(title="文件清單明細")

detail_headers = [
    ("項次", 8),
    ("原料種類", 14),
    ("廠商", 24),
    ("型號", 36),
    ("顏色", 10),
    ("檢驗 / 項目類別", 26),
    ("文件編號 / 標記", 20),
    ("特殊標註", 16),
    ("檔案名稱", 40),
    ("原廠文件下載連結 (點擊開啟)", 65)
]

ws_detail.row_dimensions[1].height = 25
ws_detail.merge_cells("A1:J1")
ws_detail["A1"] = "塑膠原料 - 原廠檢驗與證明文件清單明細 (附直接下載連結)"
ws_detail["A1"].font = font_title
ws_detail["A1"].alignment = Alignment(horizontal="left", vertical="center")

ws_detail.row_dimensions[2].height = 20
for c_idx, (h_name, w) in enumerate(detail_headers, start=1):
    col_letter = get_column_letter(c_idx)
    ws_detail.column_dimensions[col_letter].width = w
    cell = ws_detail.cell(row=2, column=c_idx, value=h_name)
    cell.font = font_h1
    cell.fill = fill_h1
    cell.alignment = align_center
    cell.border = border_header1

detail_row_idx = 3
seq_num = 1

for r in rows_data:
    cat = r['cat']
    mfg = r['mfg']
    model = r['cells']['型號'].replace('\n', ' ')
    color = r['cells']['顏色']
    
    for c in col_defs[4:]:
        cname = c['name']
        col_header = f"{c['name']} ({c['sub']})" if c['sub'] != c['name'] else c['name']
        cell_val = r['cells'].get(cname, '')
        cell_links = r['links'].get(cname, [])
        
        if not cell_val:
            continue
            
        if cell_val == 'X':
            continue # Only list actual certificates and docs
            
        if cell_links:
            for lk in cell_links:
                uri = lk['uri']
                decoded_uri = urllib.parse.unquote(uri)
                fname = os.path.basename(decoded_uri)
                lk_text = lk.get('overlapping_text', cell_val)
                
                # Check note like 含, 射出, 押出
                note = ""
                if "含" in lk_text or "含" in cell_val:
                    note = "含 (含有聲明)"
                if "射出" in lk_text or "射出" in cell_val:
                    note = "射出"
                elif "押出" in lk_text or "押出" in cell_val:
                    note = "押出"
                
                ws_detail.row_dimensions[detail_row_idx].height = 20
                ws_detail.cell(row=detail_row_idx, column=1, value=seq_num).alignment = align_center
                ws_detail.cell(row=detail_row_idx, column=2, value=cat).alignment = align_left
                ws_detail.cell(row=detail_row_idx, column=3, value=mfg).alignment = align_left
                ws_detail.cell(row=detail_row_idx, column=4, value=model).alignment = align_left
                ws_detail.cell(row=detail_row_idx, column=5, value=color).alignment = align_center
                ws_detail.cell(row=detail_row_idx, column=6, value=col_header).alignment = align_left
                ws_detail.cell(row=detail_row_idx, column=7, value=cell_val.replace('\n', ' ; ')).alignment = align_center
                ws_detail.cell(row=detail_row_idx, column=8, value=note).alignment = align_center
                ws_detail.cell(row=detail_row_idx, column=9, value=fname).alignment = align_left
                
                link_cell = ws_detail.cell(row=detail_row_idx, column=10, value=decoded_uri)
                link_cell.font = font_link
                link_cell.alignment = align_left
                link_cell.hyperlink = uri
                
                for c_i in range(1, 11):
                    ws_detail.cell(row=detail_row_idx, column=c_i).border = border_cell
                    if (seq_num % 2) == 0:
                        ws_detail.cell(row=detail_row_idx, column=c_i).fill = fill_zebra
                    if c_i != 10:
                        ws_detail.cell(row=detail_row_idx, column=c_i).font = font_data
                
                detail_row_idx += 1
                seq_num += 1
        else:
            # Document mentioned without direct hyperlink
            ws_detail.row_dimensions[detail_row_idx].height = 20
            ws_detail.cell(row=detail_row_idx, column=1, value=seq_num).alignment = align_center
            ws_detail.cell(row=detail_row_idx, column=2, value=cat).alignment = align_left
            ws_detail.cell(row=detail_row_idx, column=3, value=mfg).alignment = align_left
            ws_detail.cell(row=detail_row_idx, column=4, value=model).alignment = align_left
            ws_detail.cell(row=detail_row_idx, column=5, value=color).alignment = align_center
            ws_detail.cell(row=detail_row_idx, column=6, value=col_header).alignment = align_left
            ws_detail.cell(row=detail_row_idx, column=7, value=cell_val.replace('\n', ' ; ')).alignment = align_center
            ws_detail.cell(row=detail_row_idx, column=8, value='').alignment = align_center
            ws_detail.cell(row=detail_row_idx, column=9, value='-').alignment = align_center
            ws_detail.cell(row=detail_row_idx, column=10, value='-').alignment = align_center
            
            for c_i in range(1, 11):
                ws_detail.cell(row=detail_row_idx, column=c_i).border = border_cell
                ws_detail.cell(row=detail_row_idx, column=c_i).font = font_data
                if (seq_num % 2) == 0:
                    ws_detail.cell(row=detail_row_idx, column=c_i).fill = fill_zebra
                    
            detail_row_idx += 1
            seq_num += 1

# ==========================================
# Sheet 3: 統計與摘要 (Statistics & Notes)
# ==========================================
ws_summary = wb.create_sheet(title="統計與摘要")

ws_summary.merge_cells("A1:D1")
ws_summary["A1"] = "塑膠原料資料表 - 統計與摘要"
ws_summary["A1"].font = font_title
ws_summary.row_dimensions[1].height = 25

summary_info = [
    ("資料版本日期", "2026-08-10"),
    ("總原料大類數量", f"{len(cat_bounds)} 類 (ABS, PVC, PC, PE, PP, PBT, TPE, Copolyester, Acrylic-PC, SBC, (K), (L))"),
    ("總合作廠商數量", f"{len(mfg_bounds)} 家"),
    ("總材料型號規格數", f"{len(rows_data)} 項"),
    ("有效原廠文件/報告總數", f"{seq_num - 1} 份"),
    ("化學物質耐受性參考表連結", "http://60.251.196.100/tw/原料資料/ChemRes(化學物質耐受性參考表).pdf"),
]

for idx, (label, val) in enumerate(summary_info, start=3):
    ws_summary.row_dimensions[idx].height = 22
    c1 = ws_summary.cell(row=idx, column=1, value=label)
    c1.font = font_h2
    c1.fill = fill_h2
    c1.border = border_cell
    c1.alignment = align_left
    
    ws_summary.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=4)
    c2 = ws_summary.cell(row=idx, column=2, value=val)
    if "http" in str(val):
        c2.font = font_link
        c2.hyperlink = val
    else:
        c2.font = font_data
    c2.border = border_cell
    c2.alignment = align_left

ws_summary.column_dimensions["A"].width = 25
ws_summary.column_dimensions["B"].width = 30
ws_summary.column_dimensions["C"].width = 30
ws_summary.column_dimensions["D"].width = 30

# Save Excel files
output_dir = r"c:\Users\USER\Downloads\Project\Predictive-Material-System\新增資料夾\原料"
output_file1 = os.path.join(output_dir, "塑膠原料資料表.xlsx")
output_file2 = r"c:\Users\USER\Downloads\Project\Predictive-Material-System\塑膠原料資料表.xlsx"

wb.save(output_file1)
wb.save(output_file2)

print(f"Exported successfully to:\n1. {output_file1}\n2. {output_file2}")
